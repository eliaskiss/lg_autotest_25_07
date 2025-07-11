# Task 1
# 기존에 만든 mission.py파일을 구동시킨 상태에서
# putty가 했던 동작(Power On/Off/GetStatus)을 윈도우 프로그래밍으로 전환
# pyinstaller mission.py -> tv_monitor.exe

# Task 2
# Excel -> DB
# DB -> Excel
# file_name, _ = QFileDialog.getOpenFileName(self, 'Open file', './')

# pip install pyqt5

from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QThread, pyqtSlot, pyqtSignal

import logging
from datetime import datetime
import os
import sys

# Excel Modules
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side

# Our modules
sys.path.append('../../day_3/lec_pymysql')
sys.path.append('../../day_1/lec_pyserial')

from lec_pymysql import Database
from lec_pyserial_class import Serial

# UI 파일 포함하기
from main_ui import Ui_Dialog as Main_Ui

DISPLAY_LOG_IN_TERMNINAL = True

logger = logging.getLogger('MyLogger')
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s (%(funcName)20s:%(lineno)4d) [%(levelname)s]: %(message)s')

# Print in terminal
if DISPLAY_LOG_IN_TERMNINAL:
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

# Write in file
today = datetime.now()
today = today.strftime('%Y_%m_%d')
filename = '%s.log' % today

# If file exist, remove it
if os.path.isfile(filename):
    os.remove(filename)

file_handler = logging.FileHandler(filename)
file_handler.setFormatter(formatter)

logger.addHandler(file_handler)

COM_PORT = 'COM1'
DB_URL = '103.55.191.194'
DB_USER = 'dbuser'
DB_PW = 'dbuser'
DB_NAME = 'lg_autotest' # !! 자신의 DB이름

class MainDialog(QDialog):
    def __init__(self):
        # Display minimize, close button
        super().__init__()
        self.setWindowFlag(Qt.WindowMinimizeButtonHint, True)
        self.setWindowFlag(Qt.WindowMaximizeButtonHint, False)
        self.setWindowFlag(Qt.WindowCloseButtonHint, True)

        self.main_ui = Main_Ui()
        self.main_ui.setupUi(self)
        self.setWindowTitle('QT Sample Dialog')

        self.excel_file = None
        self.my_thread = None
        self.ser = None

        self.main_ui.btn_open_serial.clicked.connect(self.open_serial)
        self.main_ui.btn_close_serial.clicked.connect(self.close_serial)

        self.main_ui.btn_set_power.clicked.connect(self.set_power)
        self.main_ui.btn_get_power.clicked.connect(self.get_power)
        self.main_ui.btn_clear.clicked.connect(self.clear_log)

        self.main_ui.btn_find_excel.clicked.connect(self.find_excel)
        self.main_ui.btn_update_db.clicked.connect(self.update_db)
        self.main_ui.btn_export_to_excel.clicked.connect(self.export_db)
        self.main_ui.btn_exit.clicked.connect(self.close_dialog)

    def open_serial(self):
        try:
            serial_port = self.main_ui.cb_serial_port.currentText()
            self.ser = Serial()
            self.ser.openPort(serial_port)
            QMessageBox.information(self, 'Information', f'{serial_port} is opened.', QMessageBox.Ok)
            self.main_ui.gb_serial.setEnabled(True)
            self.main_ui.btn_open_serial.setEnabled(False)
            self.main_ui.btn_close_serial.setEnabled(True)
        except Exception as e:
            self.ser = None
            QMessageBox.warning(self, 'Error', str(e), QMessageBox.Ok)

    def close_serial(self):
        try:
            self.ser.closePort()
            self.ser = None
            QMessageBox.information(self, 'Information', f'Serial post is closed.', QMessageBox.Ok)
            self.main_ui.gb_serial.setEnabled(False)
            self.main_ui.btn_open_serial.setEnabled(True)
            self.main_ui.btn_close_serial.setEnabled(False)
        except Exception as e:
            self.ser = None
            QMessageBox.warning(self, 'Error', str(e), QMessageBox.Ok)

    def export_db(self):
        from_date = self.main_ui.le_date.text()
        region = self.main_ui.le_region.text()
        file_name = self.main_ui.le_export_excel.text()

        if '.xlsx' not in file_name:
            QMessageBox.warning(self, '경고', '잘못된 파일명입니다!!!', QMessageBox.Yes)
            return

        if from_date == '' or region == '':
            QMessageBox.warning(self, '경고', '설치날짜 또는 지역이 입력되지 않았습니다!!!', QMessageBox.Yes)
        else:
            self.main_ui.btn_export_to_excel.setEnabled(False)
            self.get_data_from_db(from_date, region, file_name)
            QMessageBox.information(self, '알림', '엑셀내보내기가 완료되었습니다.', QMessageBox.Yes)
            self.main_ui.btn_export_to_excel.setEnabled(True)

    @pyqtSlot(int)
    def process_init(self, init_value):
        self.main_ui.progressBar.setMaximum(init_value)

    @pyqtSlot(int)
    def process_update(self, value):
        self.main_ui.progressBar.setValue(value)

    def update_db(self):
        if self.excel_file is None:
            QMessageBox.warning(self, '경고', '선택된 파일이 없습니다!!!', QMessageBox.Yes)
        else:
            # 쓰레드없이하면 프로그램이 멈춰있음
            # self.put_data_to_db(self.excel_file)
            # self.main_ui.btn_update_db.setEnabled(False)
            # QMessageBox.information(self, '알림', 'DB 업로드가 완료되었습니다.', QMessageBox.Yes)
            # self.main_ui.btn_update_db.setEnabled(True)

            # 쓰레드
            interval = 0.1
            self.main_ui.btn_update_db.setEnabled(False)
            self.my_thread = UploadThread(self.excel_file, interval)
            self.my_thread.stopSignal.connect(self.thread_is_stopped)
            self.my_thread.progressInitSignal.connect(self.process_init)
            self.my_thread.progressUpdateSignal.connect(self.process_update)
            self.my_thread.start()


    def find_excel(self):
        file_name, _ = QFileDialog.getOpenFileName(self, 'Open file', './')

        if file_name:
            self.main_ui.le_excel_path.setText(file_name)
            self.excel_file = file_name
        else:
            QMessageBox.warning(self, '경고', '파일이 선택되지 않았습니다!!!', QMessageBox.Yes)

    def put_data_to_db(self):
        wb = load_workbook(self.excel_file_name)
        ws = wb.active

        db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
        db.connect_db()

        columns = ['station_number', 'station_name', 'region', 'address', 'latitude', 'longitude',
                   'install_date', 'lcd_count', 'qr_count', 'proc_type']

        for row in ws.iter_rows(min_row=6):
            sql = f'insert into elias_bicycle ({",".join(columns)}) values({("%s," * len(columns))[:-1]})'
            values = [_row.value for _row in row]
            db.execute_only(sql, values)

        db.commit_only()

    # todo
    def get_data_from_db(self, from_date, region, output_file_name):
        # todo : from_date 이후에 region에서 설치된 자전거 데이터를 output_file_name에 엑셀로 저장
        # ...
        wb = Workbook()
        ws = wb.active

        ws.title = '공유자전거'

        ws['A1'] = '대여소번호'
        ws.merge_cells('A1:A5')

        ws['B1'] = '보관소(대여소)명'
        ws.merge_cells('B1:B5')

        ws['C1'] = '소재자(위치)'
        ws.merge_cells('C1:F2')

        ws['C3'] = '자치구'
        ws.merge_cells('C3:C5')

        ws['D3'] = '상세주소'
        ws.merge_cells('D3:D5')

        ws['E3'] = '위도'
        ws.merge_cells('E3:E5')

        ws['F3'] = '경도'
        ws.merge_cells('F3:F5')

        ws['G1'] = '설치시기'
        ws.merge_cells('G1:G5')

        ws['H1'] = '설치형태'
        ws.merge_cells('H1:I1')

        ws['H2'] = 'LCD'
        ws.merge_cells('H2:H3')

        ws['H4'] = '거치대수'
        ws.merge_cells('H4:H5')

        ws['I2'] = 'QR'
        ws.merge_cells('I2:I3')

        ws['I4'] = '거치대수'
        ws.merge_cells('I4:I5')

        ws['J1'] = '운영방식'
        ws.merge_cells('J1:J5')

        db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
        db.connect_db()

        sql = 'SELECT * FROM elias_bicycle WHERE DATE(install_date) >= %s AND region = %s;'
        values = (from_date, region)
        data_list = db.execute_and_return(sql, values)

        for data in data_list:
            ws.append([data['station_number'], data['station_name'], data['region'],
                       data['address'], data['latitude'], data['longitude'],
                       data['install_date'], data['lcd_count'], data['qr_count'],
                       data['proc_type']])

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        bold_white_font = Font(bold=True,
                               size=12,
                               italic=None,
                               underline=None,
                               strike=None,
                               color='FFFFFF')

        background_fill = PatternFill(start_color='525E75',
                                      end_color='525E75',
                                      fill_type='solid')

        for row in ws.iter_rows(max_row=5):
            for cell in row:
                cell.fill = background_fill
                cell.font = bold_white_font

        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

        wb.save(output_file_name)

    def clear_log(self):
        self.main_ui.tb_log_power.clear()

    def set_power(self):
        # todo : rb_on 컨트롤의 상태를 체크하여 Power On(b'ka 01 01') or Power Off(b'ka 01 00') 명령어를 전송하고
        #        응답값을 받아 로그에 추가(add_log)
        is_power_on = self.main_ui.rb_on.isChecked()

        if is_power_on:
            command = b'ka 01 01\r'
        else:
            command = b'ka 01 00\r'

        self.add_log('=' * 30)
        self.add_log(f'Send Command: {command.decode()[:-1]}')
        self.ser.writePort(command)
        resp = self.ser.readLine()
        self.add_log(f'Response: {resp.decode()[:-1]}')

    def get_power(self):
        # todo : 전원상태를 가져오는 명령어(b'ka 01 ff')를 전송하여, 전원상태에 맞게 rb_on, rb_off 라디오 버튼 선택
        # ...
        command = b'ka 01 ff\r'

        self.add_log('=' * 30)
        self.add_log(f'Send Command: {command.decode()[:-1]}')
        self.ser.writePort(command)
        resp = self.ser.readLine()
        self.add_log(f'Response: {resp.decode()[:-1]}')

        # b'a 01 OK01x\r'
        # 'a 01 OK01x'
        # [7:-2] --> 01
        resp = resp[:-1].decode()
        power_state = resp[7:-1]

        if power_state == '01':
            self.main_ui.rb_on.setChecked(True)
            self.main_ui.rb_off.setChecked(False)
        else:
            self.main_ui.rb_on.setChecked(False)
            self.main_ui.rb_off.setChecked(True)

    @pyqtSlot()
    def thread_is_stopped(self):
        QMessageBox.information(self, '알림', 'DB 업로드가 완료되었습니다.', QMessageBox.Yes)
        self.main_ui.btn_update_db.setEnabled(True)

    def add_log(self, message):
        now = datetime.now()
        now = now.strftime("%H:%M:%S")
        log_message = '[%s]: %s' % (now, message)
        self.main_ui.tb_log_power.append(log_message)
        logger.info(message)

    def close_dialog(self):
        sys.exit(0)

    # ESC 무시
    def key_press_event(self, event):
        if not event.key() == Qt.Key_Escape:
            pass

class UploadThread(QThread):
    stopSignal = pyqtSignal()
    progressInitSignal = pyqtSignal(int)
    progressUpdateSignal = pyqtSignal(int)

    def __init__(self, excel_file_name, interval=10):
        super(self.__class__, self).__init__()
        self.excel_file_name = excel_file_name
        self.isRunning = True
        self.interval = interval

    def put_data_to_db(self):
        # todo: 엑셀파일(self.excel_file_name)을 읽어서, DB에 업데이트
        #       self.progressInitSignal.emit()을 이용해서 progress max값 설정
        #       self.progressUpdateSignal.emit()을 이용해서 진행상황겂 업데이트
        wb = load_workbook(self.excel_file_name)
        ws = wb.active

        db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
        db.connect_db()

        columns = ['station_number', 'station_name', 'region', 'address', 'latitude', 'longitude',
                   'install_date', 'lcd_count', 'qr_count', 'proc_type']

        total_length = len(list(ws.iter_rows(min_row=6)))
        self.progressInitSignal.emit(total_length)

        for index, row in enumerate(ws.iter_rows(min_row=6)):
            sql = f'insert into elias_bicycle ({",".join(columns)}) values({("%s," * len(columns))[:-1]})'
            values = [_row.value for _row in row]
            db.execute_only(sql, values)
            self.progressUpdateSignal.emit(index+1)

        db.commit_only()

    def run(self):
        self.put_data_to_db()
        self.stopSignal.emit()

    def stop(self):
        self.isRunning = False
        self.logSignal.emit('Thread is stopping')


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = MainDialog()
    myWindow.show()
    app.exec()