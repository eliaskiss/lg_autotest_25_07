import sys

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from icecream import ic

from lec_pymysql import Database

# Talbe Name : 자신의계정_bicycle
account = ""
table_name = f'{account}_bicycle'

DB_URL = '103.55.191.194'
DB_USER = 'dbuser'
DB_PW = 'dbuser'
DB_NAME = 'lg_autotest'

###################################################################################################
# Task 1
# public_bicycle.xlsx 파일을 읽어서, DB의 자신의계정_bicycle의 테이블에 넣기 (ex: elias_bicycle)
###################################################################################################
def put_data_to_db(excel_file_name):
    # todo: Load wb from excel file
    # ...

    # todo: Select work sheet
    # ...

    # DB 객체생성
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)

    # DB 연결
    db.connect_db()

    # Table 생성
    sql = f'CREATE TABLE if not exists {table_name} ' \
          '(`id` INT(11) NOT NULL AUTO_INCREMENT, ' \
          '`reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(), ' \
          '`station_number` INT(11) DEFAULT NULL, ' \
          '`station_name` VARCHAR(128) DEFAULT NULL, ' \
          '`region` VARCHAR(128) DEFAULT NULL, ' \
          '`address` VARCHAR(1024) DEFAULT NULL, ' \
          '`latitude` FLOAT DEFAULT NULL, ' \
          '`longitude` FLOAT DEFAULT NULL, ' \
          '`install_date` DATETIME DEFAULT NULL, ' \
          '`lcd_count` INT(11) DEFAULT NULL, ' \
          '`qr_count` INT(11) DEFAULT NULL, ' \
          '`proc_type` VARCHAR(128) DEFAULT NULL, ' \
          'KEY `id` (`id`)' \
          ') ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;'
    db.execute_and_commit(sql)

    # todo: 엑셀의 row값들을 읽어서 DB에 해당 테이블에 삽입
    # ... tip: ws.iter_rows를 사용해서 엑셀데이터를 가져와서
    #          insert 쿼리문을 사용해서 db.execute_only() db.commit_only()를 사용해서 DB에 등록

    # todo: DB연결해제
    # ...

###################################################################################################
# Task 2
# DB에 있는 자신의계정_bicycle 테이블에서 특정 데이터를 뽑아서, 엑셀로 저장하기
# ex) 2020년 이후에 서초구에 설치된 자전거 대여소 목록데이터
# sql = 'select * from elias_bicycle where date(install_date) >= "2020-01-01" and region = "서초구";'

# sql = f'select * from {table_name} where date(install_date) >= %s and region = %s;'
# from_date : "2020-01-01"
# region : "서초구"
# values = (from_date, region)
# execute_and_return(sql, values)
###################################################################################################
def get_data_from_db(from_date, region, output_file_name):
    # todo: Create new workbook
    #...

    # todo: Select Worksheet
    #...

    # todo: Rename Worksheet
    #...

    # todo: Header 생성
    #... tip: merge_cells를 사용해서 헤드설정

    # DB 객체 생성 후 연결
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
    db.connect_db()

    # todo: 조건에 맞는 데이터 가져오기
    #... tip:  f'select * from {table_name} where date(install_date) >= %s and region = %s;'문을 사용


    # todo: data_list를 가지고 엑셀의 데이터 추가
    #... tip: db로 부터 가져온 데이터를 ws.append를 사용해서 워크시트에 추가

    # todo: Excel Styling
    #... tip: Thin과 Border를 사용해서 셀의 외곽선 설정
    #         Font를 사용해서 폰트색상 및 스타일(Bold) 설정
    #         PatternFill을 사용해서 셀의 배경색 설정

    # todo: Header Style 적용
    #... tip: 위에서 생성한 PatternFill과 Font을 사용해서, fill속성과 font속성 적용

    # todo: 테두리 적용 : thin_border
    #... tip: 위에서 생성한 thin, border를 가지고 border속성에 적용

    # todo: 작성한 엑셀을 저장
    #... tip: wb.save 사용


if __name__ == '__main__':
    put_data_to_db('public_bicycle.xlsx')
    # get_data_from_db('2020-01-01', '서초구', 'new_excel.xlsx')
    pass