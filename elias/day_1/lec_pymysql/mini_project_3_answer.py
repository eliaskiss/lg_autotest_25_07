import sys

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.styles.borders import Border, Side
from icecream import ic

from lec_pymysql import Database

# Talbe Name : 자신의계정_bicycle
account = "elias"
table_name = f'{account}_bicycle'

DB_URL = '103.55.191.194'
DB_USER = 'dbuser'
DB_PW = 'dbuser'
DB_NAME = 'lg_autotest'

###################################################################################################
# Task 1
# public_bicycle.xlsx 파일을 읽어서, DB의 자신의계정_bicycle의 테이블에 넣기 (ex: elias_bicycle)
###################################################################################################
def put_data_to_db(excel_file_name):
    # Load wb from excel file
    wb = load_workbook(excel_file_name, data_only=True, read_only=True)

    # Select work sheet
    ws = wb['대여소현황']

    # DB 객체생성
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)

    # DB 연결
    db.connect_db()

    # Table 생성
    sql = f'CREATE TABLE if not exists {table_name} ' \
          '(`id` INT(11) NOT NULL AUTO_INCREMENT, ' \
          '`reg_datetime` DATETIME DEFAULT CURRENT_TIMESTAMP(), ' \
          '`station_number` INT(11) DEFAULT NULL, ' \
          '`station_name` VARCHAR(128) DEFAULT NULL, ' \
          '`region` VARCHAR(128) DEFAULT NULL, ' \
          '`address` VARCHAR(1024) DEFAULT NULL, ' \
          '`latitude` FLOAT DEFAULT NULL, ' \
          '`longitude` FLOAT DEFAULT NULL, ' \
          '`install_date` DATETIME DEFAULT NULL, ' \
          '`lcd_count` INT(11) DEFAULT NULL, ' \
          '`qr_count` INT(11) DEFAULT NULL, ' \
          '`proc_type` VARCHAR(128) DEFAULT NULL, ' \
          'KEY `id` (`id`)' \
          ') ENGINE=INNODB DEFAULT CHARSET=utf8 COLLATE=utf8_general_ci;'
    db.execute_and_commit(sql)

    # 엑셀의 row값들을 읽어서 DB에 해당 테이블에 삽입
    # ... tip: ws.iter_rows를 사용해서 엑셀데이터를 가져와서
    #          insert 쿼리문을 사용해서 db.execute_only() db.commit_only()를 사용해서 DB에 등록
    for row in ws.iter_rows(min_row=6, max_row=2591):
        # station_number = row[0].value
        # station_name = row[1].value
        # region = row[2].value
        # address = row[3].value
        # latitude = row[4].value
        # longitude = row[5].value
        # install_date = row[6].value
        # lcd_count = row[7].value
        # qr_count = row[8].value
        # proc_type = row[9].value

        # # Way I
        # sql = (f'insert into {table_name} (station_number, station_name, region, address, latitude, longitude, '
        #        f'install_date, lcd_count, qr_count, proc_type) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);')
        # values = (station_number, station_name, region, address, latitude, longitude, install_date, lcd_count, qr_count,
        #           proc_type)

        # Way II
        sql = (f'insert into {table_name} (station_number, station_name, region, address, latitude, longitude, '
               f'install_date, lcd_count, qr_count, proc_type) values(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s);')
        values = [elem.value for elem in row]
        ic(values)
        db.execute_only(sql, values)
    db.commit_only()

    # DB연결해제
    db.disconnect_db()

###################################################################################################
# Task 2
# DB에 있는 자신의계정_bicycle 테이블에서 특정 데이터를 뽑아서, 엑셀로 저장하기
# ex) 2020년 이후에 서초구에 설치된 자전거 대여소 목록데이터
# sql = 'select * from elias_bicycle where date(install_date) >= "2020-01-01" and region = "서초구";'

# sql = f'select * from {table_name} where date(install_date) >= %s and region = %s;'
# from_date : "2020-01-01"
# region : "서초구"
# values = (from_date, region)
# execute_and_return(sql, values)
###################################################################################################
def get_data_from_db(from_date, region, output_file_name):
    # Create new workbook
    wb = Workbook()

    # Select Worksheet
    ws = wb.active

    # Rename Worksheet
    ws.title = '대여소현황'

    # Header 생성
    #... tip: merge_cells를 사용해서 헤드설정
    ws['A1'] = '대여소\n번호'
    ws.merge_cells('A1:A5')

    ws['B1'] = '보관소(대여소)명'
    ws.merge_cells('B1:B5')

    ws['C1'] = '소재지(위치)'
    ws.merge_cells('C1:F2')

    ws['C3'] = '자치구'
    ws.merge_cells('C3:C5')

    ws['D3'] = '상세주소'
    ws.merge_cells('D3:D5')

    ws['E3'] = '위도'
    ws.merge_cells('E3:E5')

    ws['F3'] = '경도'
    ws.merge_cells('F3:F5')

    ws['G1'] = '설치시기'
    ws.merge_cells('G1:G5')

    ws['H1'] = '설치형태'
    ws.merge_cells('H1:I1')

    ws['H2'] = 'LCD'
    ws.merge_cells('H2:H3')

    ws['H4'] = '거치대수'
    ws.merge_cells('H4:H5')

    ws['I2'] = 'QR'
    ws.merge_cells('I2:I3')

    ws['I4'] = '거치대수'
    ws.merge_cells('I4:I5')

    ws['J1'] = '운영방식'
    ws.merge_cells('J1:J5')

    # DB 객체 생성 후 연결
    db = Database(DB_URL, DB_USER, DB_PW, DB_NAME)
    db.connect_db()

    # 조건에 맞는 데이터 가져오기
    #... tip:  f'select * from {table_name} where date(install_date) >= %s and region = %s;'문을 사용
    sql = f'select * from {table_name} where date(install_date) >= %s and region = %s;'
    values = (from_date, region)
    data_list = db.execute_and_return(sql, values)

    # data_list를 가지고 엑셀의 데이터 추가
    #... tip: db로 부터 가져온 데이터를 ws.append를 사용해서 워크시트에 추가
    for data in data_list:
        # Way I
        # ws.append([data['station_number'], data['station_name'], data['region'], data['address'],
        #            data['latitude'], data['longitude'], data['install_date'], data['lcd_count'],
        #            data['qr_count'], data['proc_type']])

        # Way II
        ws.append([elem for elem in list(data.values())[2:]])

    # Excel Styling
    #... tip: Thin과 Border를 사용해서 셀의 외곽선 설정
    #         Font를 사용해서 폰트색상 및 스타일(Bold) 설정
    #         PatternFill을 사용해서 셀의 배경색 설정
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    bold_white_font = Font(bold=True,
                           size=12,
                           italic=None,
                           underline=None,
                           strike=None,
                           color='FFFFFF')

    background_fill = PatternFill(start_color='525E75',
                                  end_color='525E75',
                                  fill_type='solid')

    # Header Style 적용
    #... tip: 위에서 생성한 PatternFill과 Font을 사용해서, fill속성과 font속성 적용
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            cell.fill = background_fill
            cell.font = bold_white_font

    # 테두리 적용 : thin_border
    #... tip: 위에서 생성한 thin, border를 가지고 border속성에 적용
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.border = thin_border

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) * 2))
    for col, value in dims.items():
        ws.column_dimensions[chr(ord('A') + (col - 1))].width = value # A, B, C, D ....

    # 작성한 엑셀을 저장
    # wb.save 사용
    wb.save(output_file_name)

if __name__ == '__main__':
    # put_data_to_db('public_bicycle.xlsx')
    get_data_from_db('2020-01-01', '서초구', 'new_excel.xlsx')